

import numpy as np
from sklearn import linear_model
from sklearn.feature_extraction import DictVectorizer
from openpyxl import *
from collections import defaultdict
from sklearn.externals import joblib
from sklearn import preprocessing
from sklearn.metrics import mean_absolute_error, mean_squared_error, median_absolute_error
import os, pickle, operator

continents = {
  "Africa": ["Algeria","Angola","Benin","Botswana","Burkina","Burundi","Cameroon","Cape Verde","Central African Republic","Chad","Comoros","Congo","Congo, Rep.","Djibouti","Egypt","Equatorial Guinea","Eritrea","Ethiopia","Gabon","Gambia","Ghana","Guinea","Guinea-Bissau","Cote d'Ivoire","Kenya","Lesotho","Liberia","Libya","Madagascar","Malawi","Mali","Mauritania","Mauritius","Morocco","Mozambique","Namibia","Niger","Nigeria","Rwanda","Sao Tome and Principe","Senegal","Seychelles","Sierra Leone","Somalia","South Africa","South Sudan","Sudan","Swaziland","Tanzania","Togo","Tunisia","Uganda","Zambia","Zimbabwe"],
  "Asia": ["Afghanistan","Bahrain","Bangladesh","Bhutan","Brunei","Myanmar","Cambodia","China","East Timor","India","Indonesia","Iran","Iraq","Israel","Japan","Jordan","Kazakhstan","North Korea","South Korea","Kuwait","Kyrgyz Republic","Lao","Lebanon","Malaysia","Maldives","Mongolia","Nepal","Oman","Pakistan","Philippines","Qatar","Russia","Saudi Arabia","Singapore","Sri Lanka","Syria","Tajikistan","Thailand","Turkey","Turkmenistan","United Arab Emirates","Uzbekistan","Vietnam","Yemen"],
  "Europe": ["Albania","Andorra","Armenia","Austria","Azerbaijan","Belarus","Belgium","Bosnia and Herzegovina","Bulgaria","Croatia","Cyprus","Czech Republic","Denmark","Estonia","Finland","France","Georgia","Germany","Greece","Hungary","Iceland","Ireland","Italy","Latvia","Liechtenstein","Lithuania","Luxembourg","Macedonia","Malta","Moldova","Monaco","Montenegro","Netherlands","Norway","Poland","Portugal","Romania","San Marino","Serbia","Slovak Republic","Slovenia","Spain","Sweden","Switzerland","Ukraine","United Kingdom","Vatican City"],
  "North America": ["Antigua and Barbuda","Bahamas","Barbados","Belize","Canada","Costa Rica","Cuba","Dominica","Dominican Republic","El Salvador","Grenada","Guatemala","Haiti","Honduras","Jamaica","Mexico","Nicaragua","Panama","Saint Kitts and Nevis","Saint Lucia","Saint Vincent and the Grenadines","Trinidad and Tobago","United States"],
  "Oceania": ["Australia","Fiji","Kiribati","Marshall Islands","Micronesia","Nauru","New Zealand","Palau","Papua New Guinea","Samoa","Solomon Islands","Tonga","Tuvalu","Vanuatu"],
  "South America": ["Argentina","Bolivia","Brazil","Chile","Colombia","Ecuador","Guyana","Paraguay","Peru","Suriname","Uruguay","Venezuela"]
}


def getRegion(countryName):
  for region in continents:
    if countryName in continents[region]:
      return region
  print "MISMATCH", countryName

# makes feature vectors from all .xlsx files in the working directory
def create_feature_vectors():
  feature_vectors = {}
  target_values = defaultdict(lambda: None)
  start_year = 1990
  end_year = 2015

  # Dictionary to store previous years' values of HIV
  previous_target_values = {} # {2: {()}
  target_year_memory_count = 3
  target_year_interval = 5
  for years_ago in range(target_year_interval, target_year_memory_count*target_year_interval+1, target_year_interval):
    previous_target_values[years_ago] = defaultdict(lambda: None)

  target_wb = load_workbook(filename = "indicator hiv estimated prevalence% 15-49.xlsx")
  target_ws = target_wb['Data']

  # store target values
  for row in target_ws.iter_rows(row_offset=1):
    country = row[0].value
    if country == None:
      break
    for cell in row:
      if cell.column != 'A':
        year = int(target_ws[cell.column + '1'].value)
        target_values[(country, year)] = float(cell.value) if cell.value is not None else None


        for yearsForward in range(target_year_interval, target_year_memory_count*target_year_interval+1, target_year_interval):
          if year + yearsForward < end_year:
            previous_target_values[yearsForward][(country, year+yearsForward)] = float(cell.value) if cell.value is not None else None

  featuresToMemoryDict = {'indicator total health expenditure perc of GDP.xlsx': (2, 3)} # : [year interval, number of years to track]

  # build feature vectors
  for filename in os.listdir(os.getcwd()):
    if filename.endswith(".xlsx") and not filename.startswith("~$") and not filename == "indicator hiv estimated prevalence% 15-49.xlsx": 
      print filename
      wb = load_workbook(filename = filename)

      feature = filename

      ws = wb['Data']

      for row in ws.iter_rows(row_offset=1):
        country = row[0].value
        if country == None:
          break
        for cell in row:
          if cell.value != None and cell.column != 'A':
            year = int(ws[cell.column + '1'].value)
            if year >= start_year:
              if (country, year) not in feature_vectors.keys():
                target = target_values[(country, year)]
                feature_vectors[(country, year)] = ({'Country': country, 'Year': year}, target)
              feature_vectors[(country, year)][0][feature] = cell.value
              
              # Storing the feature as a "past" value for future years
              if feature in featuresToMemoryDict:
                interval, count = featuresToMemoryDict[feature]
                for yearsForward in range(interval, count*interval+1, interval):
                  if year + yearsForward < end_year:
                    feature_vectors[(country, year+yearsForward)][0][feature + " " + str(yearsForward) + " years ago"] = cell.value

  for years_ago, years_ago_dict in previous_target_values.iteritems():
    for pair, prevVal in years_ago_dict.iteritems():
      if feature_vectors.get(pair):
        if prevVal != None:
            feature_vectors[pair][0]["HIV " + str(years_ago) + " years ago"] = prevVal

  return feature_vectors

feature_vectors = create_feature_vectors()
pickle.dump(feature_vectors, open('featureVectors.p', 'wb'))
#with open('featureVectors.p') as data_file:    
#    feature_vectors = pickle.load(data_file)


feature_vectors_arr = []
hiv_arr = []

for key, val in feature_vectors.iteritems():
  features, hiv = val
  if hiv is not None:
    feature_vectors_arr.append(features)
    hiv_arr.append(hiv)
  #   if cur < partition:
  #     feature_vectors_training.append(features)
  #     hiv_training.append(hiv)
  #   else:
  #     feature_vectors_test.append(features)
  #     hiv_test.append(hiv)
  # cur += 1

# print feature_vectors_test

vec = DictVectorizer()
data = vec.fit_transform(feature_vectors_arr).toarray()
feature_names = vec.get_feature_names()
# data = preprocessing.scale(data)
# hiv_arr = preprocessing.scale(hiv_arr)

partition = int(len(data) * .7)
training_data = preprocessing.scale(data[:partition])
test_data = preprocessing.scale(data[partition:])
hiv_training = preprocessing.scale(hiv_arr[:partition])
hiv_test = preprocessing.scale(hiv_arr[partition:])

# # print training_data[0]
# # print hiv_training
print "%d in training and %d in test" % (len(hiv_training), len(hiv_test))
print "Feature test len: %d  hiv test len: %d" % (len(test_data), len(hiv_test))

# print vec.get_feature_names()

# # we create 20 points
# np.random.seed(0)
# # X = np.r_[np.random.randn(10, 2) + [1, 1], np.random.randn(10, 2)]
# X = [[3, 1], [2, 1], [1, 1], [-1, -2], [-1, -3], [-4, -5]]
# y = [3, 2, 1, -1, -2, -3]
# sample_weight = 100 * np.abs(np.random.randn(20))
# # and assign a bigger weight to the last 10 samples
# sample_weight[:10] *= 10

# print X
# # print Xtest
# print y

# X = [[1, 1], [2, 2], [3, 3], [4, 4]]
# Y = [1, 2, 3, 4]
# p = [[1, 1], [2, 2]]
# e = [1, 2]

# fit the unweighted model
clf = linear_model.SGDRegressor()
clf.fit(training_data, hiv_training)
predictions = clf.predict(test_data)

# print feature_names
# print clf.coef_
weights = []
for i, coef in enumerate(clf.coef_):
  #if not "Country" in feature_names[i]:
  weights.append((feature_names[i], coef))
weights.sort(key=lambda tup: abs(tup[1]))
print weights

# print predictions

# totalError = float(0)
# for i, entry  in enumerate(predictions):
#   # print np.asscalar(entry)
#   totalError += abs(np.asscalar(entry) - hiv_test[i])
#   # print totalError

# avgError = totalError / len(predictions)
test_sd = np.std(hiv_test)
print "Sd is %f" % test_sd
print mean_absolute_error(hiv_test, predictions) * test_sd
print median_absolute_error(hiv_test, predictions) * test_sd
print mean_squared_error(hiv_test, predictions) * test_sd
print "Mean absolute error is %f" % mean_absolute_error(hiv_test, predictions) * test_sd
print "Median absolute error is %f" % median_absolute_error(hiv_test, predictions) * test_sd
print "Mean squared error is %f" % mean_squared_error(hiv_test, predictions) * test_sd

##################
# BAYESIAN NETWORK
##################

import json

from libpgm.nodedata import NodeData
from libpgm.graphskeleton import GraphSkeleton
from libpgm.discretebayesiannetwork import DiscreteBayesianNetwork
from libpgm.pgmlearner import PGMLearner

# Let's first prepare some data!

# Below would be enough, except the learner apparently doesn't work well with missing values...
# featureVectorSamples = [v[0] for v in feature_vectors.values()]
# featureVectorHIV = [v[1] for v in feature_vectors.values()]

bayes_partition = partition = int(len(feature_vectors_arr) * .7)
training_arr = feature_vectors_arr[:partition]
test_arr = feature_vectors_arr[partition:]
hiv_training_arr = hiv_arr[:partition]
hiv_test_arr = hiv_arr[partition:]

print "%d elems in training_arr out of %d total" % (len(training_arr), len(feature_vectors_arr))

vertexMap = {}

# Imputing missing values by averaging those of other countries.
# Going forward, we should totally use the linear regression code from above for it!
condensed_feature_vectors =[]
mainFeatures = ['indicator total health expenditure perc of GDP.xlsx', 'Indicator_BMI male ASM.xlsx', 'indicator food_consumption.xlsx', 'indicator_estimated incidence infectious tb per 100000.xlsx', 'indicator life_expectancy_at_birth.xlsx', 'indicator gapminder infant_mortality.xlsx']
vertices = set(mainFeatures)
for i, sample in enumerate(training_arr):
  newSample = {}
  # del sample['Country'] # This doesn't work well with discrete (and irrelevant) features like names
  # del sample['Year'] # This doesn't work well with discrete (and irrelevant) features like names

  # if feature_vectors_arr[i] != None:


# Imputing missing values by averaging those of other countries.
# Going forward, we should totally use the linear regression code from above for it!
condensed_feature_vectors_by_region = {}
mainFeatures = ['HIV 10 years ago', 'indicator total health expenditure perc of GDP.xlsx', 'Indicator_BMI male ASM.xlsx', 'indicator food_consumption.xlsx', 'indicator_estimated incidence infectious tb per 100000.xlsx', 'indicator life_expectancy_at_birth.xlsx', 'indicator gapminder infant_mortality.xlsx']
vertices = set(mainFeatures)
for i, sample in enumerate(training_arr):
  newSample = {}
  newSample['HIV'] = hiv_training_arr[i]
  
  for k in sample.keys():
    if k in vertices:
      newSample[k] = sample[k]
      if vertexMap.get(k) == None:
        vertexMap[k] = [sample]
      else:
        vertexMap[k].append(sample)
  region = getRegion(sample['Country'])
  if not region:
    break
  if region not in condensed_feature_vectors_by_region:
    condensed_feature_vectors_by_region[region] = []
  condensed_feature_vectors_by_region[region].append(newSample)


vertexAverages = {}
for vertex, samples in vertexMap.iteritems():
  numerator = 0
  for sample in samples:
    numerator += sample[vertex]
  vertexAverages[vertex] = numerator/len(samples)

for region in condensed_feature_vectors_by_region:
  for sample in condensed_feature_vectors_by_region[region]:
    for vertex in vertices:
      if vertex not in sample.keys():
        # print "Used avg for %s" % vertex
        sample[vertex] = vertexAverages[vertex]

# print condensed_feature_vectors
# Testing just 4 vertices for now (takes a really, really long time to use all of them)
# keysToRemove = list(vertices)[-1:]

# for sample in featureVectorSamples:
#   for key in keysToRemove:
#     del sample[key]

# instantiate learner 
learner = PGMLearner()

# Voila, it makes us a bayesian network!
bayesian_networks_by_region = {}
for region in condensed_feature_vectors_by_region:
  bayesian_networks_by_region[region] = learner.lg_estimatebn(condensed_feature_vectors_by_region[region])
  print region
  print json.dumps(bayesian_networks_by_region[region].Vdata, indent=2)
  print json.dumps(bayesian_networks_by_region[region].E, indent=2)

#Evaluation:
predictions = []
test_arrs_by_region = {}
hiv_test_arrs_by_region = {}
for i, sample in enumerate(test_arr):
  region = getRegion(sample['Country'])
  if not region: 
    break
  if region not in test_arrs_by_region:   
    test_arrs_by_region[region] = []
    hiv_test_arrs_by_region[region] = []
  test_arrs_by_region[region].append(sample)
  hiv_test_arrs_by_region[region].append(hiv_test_arr[i])


for region in bayesian_networks_by_region:
  print region
  predictions = []
  for sample in test_arrs_by_region[region]:
    predictionSamples = bayesian_networks_by_region[region].randomsample(500, sample)
    hivPredictionSamples = [pSample['HIV'] for pSample in predictionSamples]
    predictions.append(np.mean(hivPredictionSamples))

  # print len(hiv_test_arr)
  # print len(predictions)

  result_pairs = [(predictions[i], hiv_test_arrs_by_region[region][i]) for i in range(len(hiv_test_arrs_by_region[region]))]
  #print result_pairs

  predictions_std = preprocessing.scale(predictions)

  # With standardized predictions, SD = 1, so we're just getting how many SDs off we are
  test_sd = np.std(hiv_test_arrs_by_region[region])
  print "Sd is %f" % test_sd
  print mean_absolute_error(hiv_test_arrs_by_region[region], predictions_std) * test_sd
  print median_absolute_error(hiv_test_arrs_by_region[region], predictions_std) * test_sd
  print mean_squared_error(hiv_test_arrs_by_region[region], predictions_std) * test_sd

  # Without normalizing predictions, we need to use hiv_test_arr because that has the unnormalized HIV rates.  Now the error units are HIV rate percentage points.
  # These don't print on Matt's computer, for some reason... 
  print "Mean absolute error is %f" % mean_absolute_error(hiv_test_arrs_by_region[region], predictions)
  print "Median absolute error is %f" % median_absolute_error(hiv_test_arrs_by_region[region], predictions) 
  print "Mean squared error is %f" % mean_squared_error(hiv_test_arrs_by_region[region], predictions) 

