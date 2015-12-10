import numpy as np
from sklearn import linear_model
from sklearn.feature_extraction import DictVectorizer
from openpyxl import *
from collections import defaultdict
from sklearn.externals import joblib
from sklearn import preprocessing
from sklearn.metrics import mean_absolute_error, mean_squared_error, median_absolute_error
import os, pickle, operator
import country_mappings
import impute

 # -*- coding: utf-8 -*-

# makes feature vectors from all .xlsx files in the working directory
def create_feature_vectors():
  feature_vectors = {}
  target_values = defaultdict(lambda: None)
  start_year = 1990
  end_year = 2015
  forecast_years_ahead = 5

  # Dictionary to store previous years' values of HIV
  previous_target_values = {} # {2: {()}
  target_year_memory_count = 3
  target_year_interval = 1
  for years_ago in range(target_year_interval, target_year_memory_count*target_year_interval+1, target_year_interval):
    previous_target_values[years_ago] = defaultdict(lambda: None)

  target_wb = load_workbook(filename = "indicator hiv estimated prevalence% 15-49.xlsx")
  target_ws = target_wb['Data']

  # store target values
  for row in target_ws.iter_rows(row_offset=1):
    country = row[0].value
    if country == None:
      break
    for cell in row:
      if cell.column != 'A':
        year = int(target_ws[cell.column + '1'].value)
        target_values[(country, year-forecast_years_ahead)] = float(cell.value) if cell.value is not None else None
        for yearsForward in range(target_year_interval, target_year_memory_count*target_year_interval+1, target_year_interval):
          if year + yearsForward < end_year:
            previous_target_values[yearsForward][(country, year+yearsForward)] = float(cell.value) if cell.value is not None else None

  featuresToMemoryDict = {'indicator total health expenditure perc of GDP.xlsx': (2, 3)} # : [year interval, number of years to track]

  # build feature vectors
  for filename in os.listdir(os.getcwd()):
    if filename.endswith(".xlsx") and not filename.startswith("~$") and not filename == "indicator hiv estimated prevalence% 15-49.xlsx": 
      print filename
      wb = load_workbook(filename = filename)

      feature = filename

      ws = wb['Data']

      for row in ws.iter_rows(row_offset=1):
        country = row[0].value
        if country == None:
          break
        for cell in row:
          if cell.value != None and cell.column != 'A':
            year = int(ws[cell.column + '1'].value)
            if year >= start_year and year <= end_year - forecast_years_ahead:
              if (country, year) not in feature_vectors.keys():
                target = target_values[(country, year)]
                feature_vectors[(country, year)] = ({'Country': country, 'Year': year}, target)
              feature_vectors[(country, year)][0][feature] = cell.value
              
              # Storing the feature as a "past" value for future years
              if feature in featuresToMemoryDict:
                interval, count = featuresToMemoryDict[feature]
                for yearsForward in range(interval, count*interval+1, interval):
                  if year + yearsForward < end_year - forecast_years_ahead:
                    feature_vectors[(country, year+yearsForward)][0][feature + " " + str(yearsForward) + " years ago"] = cell.value

  for years_ago, years_ago_dict in previous_target_values.iteritems():
    for pair, prevVal in years_ago_dict.iteritems():
      if feature_vectors.get(pair):
        if prevVal != None:
            feature_vectors[pair][0]["HIV " + str(years_ago) + " years ago"] = prevVal


  # country_mappings.get_mappings(feature_vectors)

  return feature_vectors

def initialize_vectors(shouldImpute=True, useLatLong=False, includeHiv=True):
	feature_vectors = create_feature_vectors()
	if useLatLong:
		country_mappings.get_mappings(feature_vectors)
	pickle.dump(feature_vectors, open('featureVectors.p', 'wb'))
	if shouldImpute:
		impute.impute_all(includeHiv)


