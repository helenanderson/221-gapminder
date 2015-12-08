

import numpy as np
from sklearn import linear_model
from sklearn.feature_extraction import DictVectorizer
from sklearn.preprocessing import StandardScaler
from openpyxl import *
from collections import defaultdict
from sklearn.externals import joblib
import os

# makes feature vectors from all .xlsx files in the working directory
def create_feature_vectors():
  feature_vectors = {}
  target_values = defaultdict(lambda: None)
  start_year = 1990

  target_wb = load_workbook(filename = "indicator hiv estimated prevalence_ 15-49.xlsx")
  target_ws = target_wb['Data']

  # store target values
  for row in target_ws.iter_rows(row_offset=1):
    country = row[0].value
    if country == None:
      break
    for cell in row:
      if cell.column != 'A':
        year = int(target_ws[cell.column + '1'].value)
        target_values[(country, year)] = float(cell.value) if cell.value is not None else None


  # build feature vectors
  for filename in os.listdir(os.getcwd()):
    if filename.endswith(".xlsx") and not filename.startswith("~$") and not filename == "indicator hiv estimated prevalence% 15-49.xlsx": 
      print filename
      wb = load_workbook(filename = filename)

      feature = filename

      ws = wb['Data']

      for row in ws.iter_rows(row_offset=1):
        country = row[0].value
        if country == None:
          break
        for cell in row:
          if cell.value != None and cell.column != 'A':
            year = int(ws[cell.column + '1'].value)
            if year >= start_year:
              if (country, year) not in feature_vectors.keys():
                target = target_values[(country, year)]
                feature_vectors[(country, year)] = ({'Country': country, 'Year': year}, target)
              feature_vectors[(country, year)][0][feature] = cell.value

  return feature_vectors

feature_vectors = create_feature_vectors()

feature_vectors_arr = []
hiv_arr = []

for key, val in feature_vectors.iteritems():
  features, hiv = val
  if hiv is not None:
    feature_vectors_arr.append(features)
    hiv_arr.append(hiv)
  #   if cur < partition:
  #     feature_vectors_training.append(features)
  #     hiv_training.append(hiv)
  #   else:
  #     feature_vectors_test.append(features)
  #     hiv_test.append(hiv)
  # cur += 1

# print feature_vectors_test

vec = DictVectorizer()
data = vec.fit_transform(feature_vectors_arr).toarray()

partition = int(len(data) * .7)
training_data = data[:partition]
test_data = data[partition:]

# sensitive to feature scaling, so let's preprocess
scaler = StandardScaler()
scaler.fit(training_data)
#print "Training data before:\n %s" % (training_data[0])
training_data = scaler.transform(training_data)
test_data = scaler.transform(training_data)
#hiv_arr = scaler.transform(hiv_arr)

hiv_training = hiv_arr[:partition]
hiv_test = hiv_arr[partition:]

# print training_data[0]
# print hiv_training
print "%d in training and %d in test" % (len(hiv_training), len(hiv_test))
print "Feature test len: %d  hiv test len: %d" % (len(test_data), len(hiv_test))
# print vec.get_feature_names()

# # we create 20 points
# np.random.seed(0)
# # X = np.r_[np.random.randn(10, 2) + [1, 1], np.random.randn(10, 2)]
# X = [[3, 1], [2, 1], [1, 1], [-1, -2], [-1, -3], [-4, -5]]
# y = [3, 2, 1, -1, -2, -3]
# sample_weight = 100 * np.abs(np.random.randn(20))
# # and assign a bigger weight to the last 10 samples
# sample_weight[:10] *= 10

# print X
# # print Xtest
# print y




print "Training data after:\n %s \n\n\n" % (training_data[0])

# fit the unweighted model
clf = linear_model.SGDRegressor(alpha=0.01, n_iter=1)
clf.fit(training_data, hiv_training)
predictions = clf.predict(training_data)

print predictions
print hiv_test

totalError = float(0)
for i, entry  in enumerate(predictions):
  #print np.asscalar(entry)
  totalError += abs(np.asscalar(entry) - hiv_training[i])
  #print totalError

print "Average error is " % (totalError / len(predictions))
