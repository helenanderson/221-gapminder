import numpy as np
from sklearn import linear_model
from sklearn.feature_extraction import DictVectorizer
from openpyxl import *
from collections import defaultdict
from sklearn.externals import joblib
from sklearn import preprocessing
from sklearn.metrics import mean_absolute_error, mean_squared_error
import os
import json
from pprint import pprint
import pickle
from random import randint

# makes feature vectors from all .xlsx files in the working directory
def create_feature_vectors():
  feature_vectors = {}
  target_values = defaultdict(lambda: None)
  start_year = 1990

  target_wb = load_workbook(filename = "indicator hiv estimated prevalence_ 15-49.xlsx")
  target_ws = target_wb['Data']

  # store target values
  for row in target_ws.iter_rows(row_offset=1):
    country = row[0].value
    if country == None:
      break
    for cell in row:
      if cell.column != 'A':
        year = int(target_ws[cell.column + '1'].value)
        target_values[(country, year)] = float(cell.value) if cell.value is not None else None


  # build feature vectors
  for filename in os.listdir(os.getcwd()):
    if filename.endswith(".xlsx") and not filename.startswith("~$") and not filename == "indicator hiv estimated prevalence_ 15-49.xlsx": 
      print filename
      wb = load_workbook(filename = filename)

      feature = filename

      ws = wb['Data']

      for row in ws.iter_rows(row_offset=1):
        country = row[0].value
        if country == None:
          break
        for cell in row:
          if cell.value != None and cell.column != 'A':
            year = int(ws[cell.column + '1'].value)
            if year >= start_year:
              if (country, year) not in feature_vectors.keys():
                target = target_values[(country, year)]
                feature_vectors[(country, year)] = ({'Country': country, 'Year': year}, target)
              feature_vectors[(country, year)][0][feature] = cell.value
  
  pickle.dump(feature_vectors, open('featureVectors.p', 'wb'))
  return feature_vectors

# feature_vectors = create_feature_vectors()
# Saving time by replacing the above line with the below:
with open('featureVectors.p') as data_file:    
    feature_vectors = pickle.load(data_file)
#pprint(feature_vectors)

feature_vectors_arr = []
hiv_arr = []

for key, val in feature_vectors.iteritems():
  features, hiv = val
  if hiv is not None:
    feature_vectors_arr.append(features)
    hiv_arr.append(hiv)
  #   if cur < partition:
  #     feature_vectors_training.append(features)
  #     hiv_training.append(hiv)
  #   else:
  #     feature_vectors_test.append(features)
  #     hiv_test.append(hiv)
  # cur += 1

# print feature_vectors_test

vec = DictVectorizer()
data = vec.fit_transform(feature_vectors_arr).toarray()
# data = preprocessing.scale(data)
# hiv_arr = preprocessing.scale(hiv_arr)

partition = int(len(data) * .7)
training_data = preprocessing.scale(data[:partition])
test_data = preprocessing.scale(data[partition:])
hiv_training = preprocessing.scale(hiv_arr[:partition])
hiv_test = preprocessing.scale(hiv_arr[partition:])

# # print training_data[0]
# # print hiv_training
print "%d in training and %d in test" % (len(hiv_training), len(hiv_test))
print "Feature test len: %d  hiv test len: %d" % (len(test_data), len(hiv_test))
# print vec.get_feature_names()

# # we create 20 points
# np.random.seed(0)
# # X = np.r_[np.random.randn(10, 2) + [1, 1], np.random.randn(10, 2)]
# X = [[3, 1], [2, 1], [1, 1], [-1, -2], [-1, -3], [-4, -5]]
# y = [3, 2, 1, -1, -2, -3]
# sample_weight = 100 * np.abs(np.random.randn(20))
# # and assign a bigger weight to the last 10 samples
# sample_weight[:10] *= 10

# print X
# # print Xtest
# print y

# X = [[1, 1], [2, 2], [3, 3], [4, 4]]
# Y = [1, 2, 3, 4]
# p = [[1, 1], [2, 2]]
# e = [1, 2]

# fit the unweighted model
clf = linear_model.SGDRegressor(alpha=0.01, n_iter=100)
clf.fit(training_data, hiv_training)
predictions = clf.predict(test_data)

# print predictions


# totalError = float(0)
# for i, entry  in enumerate(predictions):
#   # print np.asscalar(entry)
#   totalError += abs(np.asscalar(entry) - hiv_test[i])
#   # print totalError

# avgError = totalError / len(predictions)

print "Mean absolute error is %f" % mean_absolute_error(hiv_test, predictions)
print "Mean squared error is %f" % mean_squared_error(hiv_test, predictions)


##################
# BAYESIAN NETWORK
##################

import json

from libpgm.nodedata import NodeData
from libpgm.graphskeleton import GraphSkeleton
from libpgm.discretebayesiannetwork import DiscreteBayesianNetwork
from libpgm.pgmlearner import PGMLearner

# Let's first prepare some data!

# Below would be enough, except the learner apparently doesn't work well with missing values...
featureVectorSamples = [v[0] for v in feature_vectors.values()]
featureVectorHIV = [v[1] for v in feature_vectors.values()]

vertexMap = {}

# Imputing missing values by averaging those of other countries.
# Going forward, we should totally use the linear regression code from above for it!
vertices = set()
for i, sample in enumerate(featureVectorSamples):
  del sample['Country'] # This doesn't work well with discrete (and irrelevant) features like names
  del sample['Year'] # This doesn't work well with discrete (and irrelevant) features like names

  if featureVectorHIV[i] != None:
    sample['HIV'] = featureVectorHIV[i]
  for k in sample.keys():
    vertices.add(k)
    if vertexMap.get(k) == None:
      vertexMap[k] = [sample]
    else:
      vertexMap[k].append(sample)

vertexAverages = {}
for vertex, samples in vertexMap.iteritems():
  numerator = 0
  for sample in samples:
    numerator += sample[vertex]
  vertexAverages[vertex] = numerator/len(samples)

for sample in featureVectorSamples:
  for vertex in vertices:
    if vertex not in sample.keys():
      sample[vertex] = vertexAverages[vertex]


# Testing just 4 vertices for now (takes a really, really long time to use all of them)
keysToRemove = list(vertices)[5:]
#keysToRemove.remove('HIV')

for sample in featureVectorSamples:
  for key in keysToRemove:
    del sample[key]

# instantiate learner 
learner = PGMLearner()

# Voila, it makes us a bayesian network!
result = learner.lg_estimatebn(featureVectorSamples, pvalparam = 0.10)

# output
print json.dumps(result.Vdata, indent=2)
print json.dumps(result.E, indent=2)

# For progress report: previous things we tried!

# Hackily removes all vertices with missing values, leaving just country name and year :P
# Instead, we should totally impute values using our linear classifier!
# commonVertices = vertices
# for sample in featureVectorSamples:
#   commonVertices2 = set([v for v in commonVertices])
#   for v in commonVertices:
#     if v not in sample.keys():
#       commonVertices2.remove(v)
#   commonVertices = set(commonVertices2)

# for i in range(len(featureVectorSamples)):
#   # for k in featureVectorSamples[i].keys():
#   #   if k not in commonVertices:
#   #     del featureVectorSamples[i][k]
#   for k, v in featureVectorSamples[i].iteritems():
#     if isinstance(v, unicode):
#       featureVectorSamples[i][k] = featureVectorSamples[i][k].encode('ascii', errors='backslashreplace')
# pprint(featureVectorSamples)
# print "Replaced features"

# Makes a fully connected (i.e. every vertex connects to every other) graph
# def createGraphSkeletonFile(nodeData):
#   skeleton = {'V':[], 'E':[]}
#   vertices = set()
#   print nodeData
#   for n in nodeData:
#     for k, v in n.iteritems():
#       vertices.add(k)

#   vertexList = list(vertices)
#   skeleton['V'] = vertexList
#   for v1 in vertices:
#     for v2 in vertices:
#       if not v1==v2:
#         skeleton['E'].append([v1, v2])

#   f = open('graphSkeleton.txt', 'w')
#   json.dump(skeleton, f)



# Randomized but not-so-hot imputation:
# i = randint(0,len(vertexMap[v])-1)      
# sample[vertex] = vertexMap[vertex][i][vertex]
